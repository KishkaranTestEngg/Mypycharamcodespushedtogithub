import openpyxl
class ExcelData:
    def __init__(self,filename,sheetname):
        self.fname = filename
        self.sheetname = sheetname

    def load_excel_workbook(self):
        work_book = openpyxl.load_workbook(filename=self.fname)
        work_sheet = work_book[self.sheetname]
        return work_book, work_sheet

    def rows_count(self,worksheet):
        rows = worksheet.max_row
        return rows

    def colmn_count(self,worksheet):
        colmn = worksheet.max_column
        return colmn

    def get_data_(self,worksheet):
        self.rows = self.rows_count(worksheet)
        self.columns = self.colmn_count(worksheet)
        user_data = []
        for each_row in range(2,self.rows+1):
            for each_col in range(1,self.columns+1):
                cell_value = worksheet.cell(row=each_row, column=each_col).value
                user_data.append(cell_value)
        return user_data

    def save_workbook(self,workbook):
        workbook.save(self.fname)

    def close_workbook(self,workbook):
        workbook.close()

